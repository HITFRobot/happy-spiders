# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import load_workbook
import os

data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../data')


class ThreeGorgesPipeline(object):
    def process_item(self, item, spider):
        return item


class SanXiaExcelPipeline(object):
    def __init__(self) -> None:
        self.file = os.path.join(data_dir, 'sanxia.xlsx')
        self.excel = load_workbook(self.file)

    def process_item(self, item, spider):
        if item['name'] != 'sanxia':
            return item
        year = item['time'].split('-')[0]
        # 获取指定的sheet
        sheet = self.excel.get_sheet_by_name(year)
        # 准备插入数据
        rows = []
        date = item['time'].split('-', maxsplit=1)[1]
        # 02
        row02 = []
        row02.append(date)
        row02.append('2:00')
        row02.append(item['rk02'])
        row02.append(item['ck02'])
        row02.append(item['sy02'])
        row02.append(item['xy02'])
        rows.append(row02)
        # 08
        row08 = []
        row08.append(date)
        row08.append('8:00')
        row08.append(item['rk08'])
        row08.append(item['ck08'])
        row08.append(item['sy08'])
        row08.append(item['xy08'])
        rows.append(row08)
        # 14
        row14 = []
        row14.append(date)
        row14.append('14:00')
        row14.append(item['rk14'])
        row14.append(item['ck14'])
        row14.append(item['sy14'])
        row14.append(item['xy14'])
        rows.append(row14)
        # 20
        row20 = []
        row20.append(date)
        row20.append('20:00')
        row20.append(item['rk20'])
        row20.append(item['ck20'])
        row20.append(item['sy20'])
        row20.append(item['xy20'])
        rows.append(row14)
        # 插入
        sheet.append(row02)
        sheet.append(row08)
        sheet.append(row14)
        sheet.append(row20)
        self.excel.save(self.file)
        return item

    def spider_closed(self, spider):
        self.excel.close()


class XiangJiaExcelPipeline(object):
    def __init__(self) -> None:
        self.file = os.path.join(data_dir, 'xiangjiaba.xlsx')
        self.excel = load_workbook(self.file)

    def process_item(self, item, spider):
        if item['name'] != 'xiangjia':
            return item
        year = item['time'].split('-')[0]
        # 获取指定的sheet
        sheet = self.excel.get_sheet_by_name(year)
        # 准备插入数据
        rows = []
        date = item['time'].split('-', maxsplit=1)[1]
        # 02
        row02 = []
        row02.append(date)
        row02.append('2:00')
        row02.append(item['rk02'])
        row02.append(item['ck02'])
        row02.append(item['sy02'])
        row02.append(item['xy02'])
        rows.append(row02)
        # 08
        row08 = []
        row08.append(date)
        row08.append('8:00')
        row08.append(item['rk08'])
        row08.append(item['ck08'])
        row08.append(item['sy08'])
        row08.append(item['xy08'])
        rows.append(row08)
        # 14
        row14 = []
        row14.append(date)
        row14.append('14:00')
        row14.append(item['rk14'])
        row14.append(item['ck14'])
        row14.append(item['sy14'])
        row14.append(item['xy14'])
        rows.append(row14)
        # 20
        row20 = []
        row20.append(date)
        row20.append('20:00')
        row20.append(item['rk20'])
        row20.append(item['ck20'])
        row20.append(item['sy20'])
        row20.append(item['xy20'])
        rows.append(row14)
        # 插入
        sheet.append(row02)
        sheet.append(row08)
        sheet.append(row14)
        sheet.append(row20)
        self.excel.save(self.file)
        return item

    def spider_closed(self, spider):
        self.excel.close()
